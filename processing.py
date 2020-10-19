import over_stats
from openpyxl import Workbook
from openpyxl import load_workbook as loadBook

loc = 'C:\\Users\\schle\\Documents\\PMAOWL\\stats.xlsx'

players = {}
def convertTime(time):
    
    temp = time.split(':', 3)

    if len(temp) == 2:
        for x in range(0,2):
            temp[x] = float(temp[x])
    else:
        for x in range(0,3):
            temp[x] = float(temp[x])

    timeCon = 0

    if len(temp) == 2:
        for x in range(0,2):
            timeCon = (temp[0]/60) + (temp[1]/(60*60))
    else:
        for x in range(0,3):
            timeCon = (temp[0]) + (temp[1]/60) + (temp[2]/(60*60))

    return timeCon

def findKey(mydict, search):
    tempDict = {}
    for x in mydict:
        tempDict[mydict[x]] = x
    
    return tempDict[search]

def find3MostPlayed(battletag):
    mostPlayed = {'hero1': {}, 'hero2': {}, 'hero3': {}}
    playerData = over_stats.PlayerProfile(battletag)
    heroes = playerData.stat_heroes('competitive')
    heroesDict = {}
    for x in heroes:
        if((x != 'ALL HEROES')):
            heroesDict[x] = convertTime(playerData.stats('competitive', x, 'Game', 'Time Played'))

    for x in range(0,3):
        key = findKey(heroesDict, max(heroesDict.values()))
        wins = playerData.stats('competitive', key, 'Game', 'Games Won')
        total = playerData.stats('competitive', key, 'Game', 'Games Played')
        mostPlayed['hero'+str(x+1)] = [key, heroesDict[key], (wins/total)]
        heroesDict.pop(key)

    mostPlayed['totalPlaytime'] = convertTime(playerData.stats('competitive', 'ALL HEROES', 'Game', 'Time Played'))
    mostPlayed['totalWinrate'] = (playerData.stats('competitive', 'ALL HEROES', 'Game', 'Games Won'))/(playerData.stats('competitive', 'ALL HEROES', 'Game', 'Games Played'))

    print(mostPlayed)

    return mostPlayed

def initializeDict(players):
    nameCol, tankCol, dpsCol, supCol, teamCol, divCol = 'A', 'B', 'C', 'D', 'E', 'F'
    dataCol, hero1Col, hero1PlayCol, hero1WinCol = 'G', 'H', 'I', 'J'
    hero2Col, hero2PlayCol, hero2WinCol = 'K', 'L', 'M'
    hero3Col, hero3PlayCol, hero3WinCol = 'N', 'O', 'P'
    totalPlayCol, totalWinCol = 'Q', 'R'
    wb = loadBook(loc)
    sheets = wb.sheetnames
    ws = wb[sheets[0]]

    rows= ws.max_row
    
    for rowNum in range(2, (rows+1)):
        if(ws[dataCol+str(rowNum)].value == 'T'):
            players[ws[nameCol+str(rowNum)].value] = {'tank': (ws[tankCol+str(rowNum)].value), 
            'dps': (ws[dpsCol+str(rowNum)].value), 'sup': (ws[supCol+str(rowNum)].value), 
            'team': (ws[teamCol+str(rowNum)].value), 'div': (ws[divCol+str(rowNum)].value),
            'data': (ws[dataCol+str(rowNum)].value), 'hero1': (ws[hero1Col+str(rowNum)].value),
            'hero1Play': (ws[hero1PlayCol+str(rowNum)].value), 'hero1Win': (ws[hero1WinCol+str(rowNum)].value),
            'hero2': (ws[hero2Col+str(rowNum)].value), 'hero2Play': (ws[hero2PlayCol+str(rowNum)].value),
            'hero2Win': (ws[hero2WinCol+str(rowNum)].value), 'hero3': (ws[hero3Col+str(rowNum)].value),
            'hero3Play': (ws[hero3PlayCol+str(rowNum)].value), 'hero3Win': (ws[hero3WinCol+str(rowNum)].value),
            'totalPlay': (ws[totalPlayCol+str(rowNum)].value), 'totalWin': (ws[totalWinCol+str(rowNum)].value)}
        else:
            
            print(ws[nameCol+str(rowNum)].value)
            mostPlayed = find3MostPlayed(ws[nameCol+str(rowNum)].value)

            ws[dataCol+str(rowNum)].value = 'T'
            ws[hero1Col+str(rowNum)].value = mostPlayed['hero1'][0]
            ws[hero1PlayCol+str(rowNum)].value = mostPlayed['hero1'][1]
            ws[hero1WinCol+str(rowNum)].value = mostPlayed['hero1'][2]

            ws[hero2Col+str(rowNum)].value = mostPlayed['hero2'][0]
            ws[hero2PlayCol+str(rowNum)].value = mostPlayed['hero2'][1]
            ws[hero2WinCol+str(rowNum)].value = mostPlayed['hero2'][2]

            ws[hero3Col+str(rowNum)].value = mostPlayed['hero3'][0]
            ws[hero3PlayCol+str(rowNum)].value = mostPlayed['hero3'][1]
            ws[hero3WinCol+str(rowNum)].value = mostPlayed['hero3'][2]

            ws[totalPlayCol+str(rowNum)].value = mostPlayed['totalPlaytime']
            ws[totalWinCol+str(rowNum)].value = mostPlayed['totalWinrate']
            wb.save(loc)

            players[ws[nameCol+str(rowNum)].value] = {'tank': (ws[tankCol+str(rowNum)].value), 
            'dps': (ws[dpsCol+str(rowNum)].value), 'sup': (ws[supCol+str(rowNum)].value), 
            'team': (ws[teamCol+str(rowNum)].value), 'div': (ws[divCol+str(rowNum)].value),
            'data': (ws[dataCol+str(rowNum)].value), 'hero1': (ws[hero1Col+str(rowNum)].value),
            'hero1Play': (ws[hero1PlayCol+str(rowNum)].value), 'hero1Win': (ws[hero1WinCol+str(rowNum)].value),
            'hero2': (ws[hero2Col+str(rowNum)].value), 'hero2Play': (ws[hero2PlayCol+str(rowNum)].value),
            'hero2Win': (ws[hero2WinCol+str(rowNum)].value), 'hero3': (ws[hero3Col+str(rowNum)].value),
            'hero3Play': (ws[hero3PlayCol+str(rowNum)].value), 'hero3Win': (ws[hero3WinCol+str(rowNum)].value),
            'totalPlay': (ws[totalPlayCol+str(rowNum)].value), 'totalWin': (ws[totalWinCol+str(rowNum)].value)}
    


initializeDict(players)

print(players)